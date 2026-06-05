"""산출물 내보내기 라우터 — ERD 메타로부터 문서(엑셀 등)를 생성한다.

클라이언트(ENTITIES)가 ERD 를 JSON 으로 보내면 사이드카가 파일을 만들어 반환한다.
데스크탑 사이드카 전용(openpyxl 사용 — 표준사전 엑셀과 동일 의존).
"""
import io
from typing import List, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

router = APIRouter()


class SpecColumn(BaseModel):
    logicalName: Optional[str] = ""
    physicalName: Optional[str] = ""
    type: Optional[str] = ""
    kind: Optional[str] = ""          # pk | fk | normal
    notNull: Optional[bool] = False
    defaultValue: Optional[str] = ""
    description: Optional[str] = ""


class SpecTable(BaseModel):
    logicalName: Optional[str] = ""
    physicalName: Optional[str] = ""
    description: Optional[str] = ""
    columns: List[SpecColumn] = []


class TableSpecBody(BaseModel):
    title: Optional[str] = "테이블 정의서"
    tables: List[SpecTable] = []


@router.post("/table-spec")
def table_spec_xlsx(body: TableSpecBody):
    """ERD 테이블 목록 → 엑셀 테이블 정의서(.xlsx). 목차 + 테이블정의서 2시트."""
    if not body.tables:
        raise HTTPException(status_code=400, detail="tables 가 비어 있습니다.")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치 — 사이드카 재빌드 필요")

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()

    # ── 시트1: 목차 ──
    idx = wb.active
    idx.title = "목차"
    idx.append(["순번", "논리명", "물리명", "컬럼수", "설명"])
    for c in range(1, 6):
        cell = idx.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, border
    for i, t in enumerate(body.tables, 1):
        idx.append([i, t.logicalName or "", t.physicalName or "", len(t.columns), t.description or ""])
        for c in range(1, 6):
            idx.cell(row=i + 1, column=c).border = border
    for col, w in zip("ABCDE", [6, 26, 26, 8, 50]):
        idx.column_dimensions[col].width = w

    # ── 시트2: 테이블정의서 ──
    ws = wb.create_sheet("테이블정의서")
    headers = ["순번", "논리명", "물리명", "데이터타입", "PK", "FK", "NN", "기본값", "설명"]
    for col, w in zip("ABCDEFGHI", [6, 24, 24, 18, 5, 5, 5, 14, 40]):
        ws.column_dimensions[col].width = w

    r = 1
    for ti, t in enumerate(body.tables, 1):
        ws.cell(row=r, column=1, value=f"{ti}. {t.logicalName or ''}  [{t.physicalName or ''}]")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        tc = ws.cell(row=r, column=1)
        tc.fill, tc.font, tc.alignment = title_fill, title_font, left
        r += 1
        if t.description:
            ws.cell(row=r, column=1, value="설명: " + t.description)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            ws.cell(row=r, column=1).alignment = left
            r += 1
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, border
        r += 1
        for i, col in enumerate(t.columns, 1):
            vals = [i, col.logicalName or "", col.physicalName or "", col.type or "",
                    "●" if col.kind == "pk" else "", "●" if col.kind == "fk" else "",
                    "●" if col.notNull else "", col.defaultValue or "", col.description or ""]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.border = border
                cell.alignment = center if ci in (1, 5, 6, 7) else left
            r += 1
        r += 1   # 테이블 간 빈 줄

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=table_spec.xlsx"},
    )
