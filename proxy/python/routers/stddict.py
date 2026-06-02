# routers/stddict.py
# 표준사전(word/domain/term) — 사이드카가 sqlite 파일을 직접 소유·CRUD.
# 저장 위치: 시스템 DB aerm_storage (~/.uxermanager/aerm_storage.db) 의 word/domain/term 테이블
# 엑셀 import 는 tools/build_std_sqlite.py 의 시트→테이블 매핑을 그대로 따른다.
import datetime
import io
import sqlite3
import tempfile
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db.system_db import connect as _system_connect, SYSTEM_DB_FILE, DATA_DIR

router = APIRouter()

# ── 테이블/컬럼 화이트리스트 (js/std_dict.js STD_COLS 와 동기화) ──────────────
STD_COLS = {
    "word": [
        "name", "abbr", "full_name", "descr", "domain_class",
        "format_word", "synonym", "forbidden", "revision", "org",
        "approved", "reg_user", "reg_at", "upd_user", "upd_at",
    ],
    "domain": [
        "class_name", "name", "group_name", "descr", "data_type",
        "len", "scale", "store_fmt", "disp_fmt", "unit", "allowed",
        "revision", "org", "approved", "reg_user", "reg_at",
    ],
    "term": [
        "name", "abbr", "descr", "domain_name", "allowed",
        "store_fmt", "disp_fmt", "gov_code_name", "gov_org", "synonym",
        "revision", "org", "approved", "reg_user", "reg_at",
    ],
}

AUDIT_COLS = {"reg_user", "reg_at", "upd_user", "upd_at"}

SEARCH_COLS = {
    "word": ["name", "abbr", "descr"],
    "domain": ["name", "class_name", "descr"],
    "term": ["name", "abbr", "descr"],
}

# ── 엑셀 시트 → 테이블 매핑 (tools/build_std_sqlite.py 와 동일) ───────────────
SHEETS = {
    "표준단어": ("word", [
        ("표준단어명", "name"), ("영문약어", "abbr"), ("영문전체명", "full_name"),
        ("설명", "descr"), ("도메인분류", "domain_class"), ("형식단어", "format_word"),
        ("이음동의어", "synonym"), ("금칙어", "forbidden"), ("개정구분", "revision"),
        ("조직구분", "org"), ("승인여부", "approved"), ("등록자", "reg_user"),
        ("등록일시", "reg_at"), ("수정자", "upd_user"), ("수정일시", "upd_at"),
    ]),
    "표준도메인": ("domain", [
        ("도메인분류명", "class_name"), ("도메인명", "name"), ("도메인그룹명", "group_name"),
        ("설명", "descr"), ("데이터타입", "data_type"), ("길이", "len"),
        ("소수점", "scale"), ("저장형식", "store_fmt"), ("표현형식", "disp_fmt"),
        ("단위", "unit"), ("허용값", "allowed"), ("개정구분", "revision"),
        ("조직구분", "org"), ("승인여부", "approved"), ("등록자", "reg_user"),
        ("등록일시", "reg_at"),
    ]),
    "표준용어": ("term", [
        ("표준용어명", "name"), ("영문약어", "abbr"), ("설명", "descr"),
        ("도메인명", "domain_name"), ("허용값", "allowed"), ("저장형식", "store_fmt"),
        ("표현형식", "disp_fmt"), ("행정표준코드명", "gov_code_name"), ("소관기관명", "gov_org"),
        ("이음동의어", "synonym"), ("개정구분", "revision"), ("조직구분", "org"),
        ("승인여부", "approved"), ("등록자", "reg_user"), ("등록일시", "reg_at"),
    ]),
}


# ── 공통 헬퍼 ────────────────────────────────────────────────────────────────
def _now() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _check_table(table: str):
    if table not in STD_COLS:
        raise HTTPException(status_code=400, detail=f"알 수 없는 테이블: {table}")


def _conn() -> sqlite3.Connection:
    # 표준사전은 시스템 DB(aerm_storage)의 word/domain/term 테이블을 사용한다.
    return _system_connect()


def _create_schema(con: sqlite3.Connection):
    """3테이블이 없으면 빈 스키마 생성 (id + TEXT 컬럼)."""
    for table, cols in STD_COLS.items():
        col_defs = "id INTEGER PRIMARY KEY, " + ", ".join(f"{c} TEXT" for c in cols)
        con.execute(f"CREATE TABLE IF NOT EXISTS {table} ({col_defs})")


def _create_indexes(con: sqlite3.Connection):
    con.execute("CREATE INDEX IF NOT EXISTS idx_word_name   ON word(name)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_word_abbr   ON word(abbr)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_domain_name ON domain(name)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_term_name   ON term(name)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_term_abbr   ON term(abbr)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_term_domain ON term(domain_name)")


def _ensure_db():
    """작업본 파일이 없으면 빈 스키마로 생성."""
    con = _conn()
    try:
        _create_schema(con)
        _create_indexes(con)
        con.commit()
    finally:
        con.close()


def _row_to_dict(row: sqlite3.Row) -> dict:
    return {k: row[k] for k in row.keys()}


# ── GET /stddict/status ──────────────────────────────────────────────────────
@router.get("/status")
def status():
    counts = {}
    con = _conn()
    try:
        _create_schema(con)  # 시스템 DB에 표준사전 테이블 없으면 생성
        for table in STD_COLS:
            counts[table] = con.execute(f"SELECT COUNT(*) AS c FROM {table}").fetchone()["c"]
    finally:
        con.close()
    # 표준사전 데이터가 한 건이라도 있으면 초기화 완료로 간주 (시드 주입 여부 판단용)
    initialized = any(v > 0 for v in counts.values())
    return {"initialized": initialized, "counts": counts}


# ── GET /stddict/list ────────────────────────────────────────────────────────
@router.get("/list")
def list_rows(
    table: str = Query(...),
    q: str = Query(""),
    onlyApproved: bool = Query(False),
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0),
):
    _check_table(table)
    _ensure_db()
    con = _conn()
    try:
        conds, params = [], []
        kw = (q or "").strip()
        if kw:
            like = "%" + kw + "%"
            clauses = " OR ".join(f"{c} LIKE ?" for c in SEARCH_COLS[table])
            conds.append(f"({clauses})")
            params.extend([like] * len(SEARCH_COLS[table]))
        if onlyApproved:
            conds.append("approved='Y'")
        where = ("WHERE " + " AND ".join(conds)) if conds else ""

        total = con.execute(f"SELECT COUNT(*) AS c FROM {table} {where}", params).fetchone()["c"]
        rows = con.execute(
            f"SELECT * FROM {table} {where} ORDER BY id LIMIT ? OFFSET ?",
            [*params, limit, offset],
        ).fetchall()
        return {"rows": [_row_to_dict(r) for r in rows], "total": total}
    finally:
        con.close()


# ── GET /stddict/index ───────────────────────────────────────────────────────
# 자동완성용 경량 인덱스 — 한 테이블의 (name, abbr)만 전부 반환.
# 프론트가 세션당 1회 로드해 클라이언트에서 필터한다(키 입력마다 요청 방지).
@router.get("/index")
def get_index(table: str = Query("term")):
    _check_table(table)
    _ensure_db()
    has_abbr = "abbr" in STD_COLS[table]
    con = _conn()
    try:
        cols = "name, abbr" if has_abbr else "name"
        rows = con.execute(f"SELECT {cols} FROM {table} ORDER BY name").fetchall()
        items = [{"name": r["name"], "abbr": (r["abbr"] if has_abbr else None)} for r in rows]
        return {"items": items}
    finally:
        con.close()


# ── POST /stddict/row ────────────────────────────────────────────────────────
class RowBody(BaseModel):
    table: str
    values: dict


@router.post("/row")
def insert_row(body: RowBody):
    _check_table(body.table)
    _ensure_db()
    allowed = STD_COLS[body.table]
    merged = dict(body.values or {})

    now, user = _now(), "user"
    merged["reg_user"] = merged.get("reg_user") or user
    merged["reg_at"] = merged.get("reg_at") or now
    if "upd_user" in allowed:
        merged["upd_user"] = merged.get("upd_user") or user
    if "upd_at" in allowed:
        merged["upd_at"] = merged.get("upd_at") or now
    if not merged.get("approved"):
        merged["approved"] = "미승인"

    cols = [c for c in allowed if c in merged]
    vals = [merged.get(c) for c in cols]
    sql = f"INSERT INTO {body.table} ({','.join(cols)}) VALUES ({','.join('?' * len(cols))})"

    con = _conn()
    try:
        cur = con.execute(sql, vals)
        con.commit()
        return {"ok": True, "id": cur.lastrowid}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"삽입 오류: {e}")
    finally:
        con.close()


# ── PUT /stddict/row/{row_id} ────────────────────────────────────────────────
@router.put("/row/{row_id}")
def update_row(row_id: int, body: RowBody):
    _check_table(body.table)
    _ensure_db()
    allowed = STD_COLS[body.table]
    merged = dict(body.values or {})

    # reg_* 보존, upd_* 자동 세팅
    merged.pop("reg_user", None)
    merged.pop("reg_at", None)
    if "upd_user" in allowed:
        merged["upd_user"] = "user"
    if "upd_at" in allowed:
        merged["upd_at"] = _now()

    cols = [c for c in allowed if c in merged]
    if not cols:
        return {"ok": False, "detail": "변경할 컬럼 없음"}
    sets = ",".join(f"{c}=?" for c in cols)
    vals = [merged.get(c) for c in cols] + [row_id]

    con = _conn()
    try:
        cur = con.execute(f"UPDATE {body.table} SET {sets} WHERE id=?", vals)
        con.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="행을 찾을 수 없습니다.")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"수정 오류: {e}")
    finally:
        con.close()


# ── DELETE /stddict/row/{row_id} ─────────────────────────────────────────────
@router.delete("/row/{row_id}")
def delete_row(row_id: int, table: str = Query(...)):
    _check_table(table)
    _ensure_db()
    con = _conn()
    try:
        cur = con.execute(f"DELETE FROM {table} WHERE id=?", [row_id])
        con.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="행을 찾을 수 없습니다.")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"삭제 오류: {e}")
    finally:
        con.close()


# ── POST /stddict/import-excel ───────────────────────────────────────────────
def _val(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        if isinstance(v, datetime.datetime):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return v.strftime("%Y-%m-%d")
    if v is None:
        return None
    return str(v) if not isinstance(v, (int, float)) else v


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...)):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치 — 사이드카 재빌드 필요")

    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 열기 실패: {e}")

    _ensure_db()
    con = _conn()
    counts = {}
    try:
        for sheet_name, (table, col_map) in SHEETS.items():
            if sheet_name not in wb.sheetnames:
                counts[table] = -1  # 시트 없음
                continue
            ws = wb[sheet_name]
            eng_cols = [eng for _, eng in col_map]

            # 전체 교체: 기존 행 삭제 후 재삽입
            con.execute(f"DELETE FROM {table}")

            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                counts[table] = 0
                continue
            header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
            col_indices = [header_idx.get(han) for han, _ in col_map]

            placeholders = ",".join("?" * len(eng_cols))
            sql_ins = f"INSERT INTO {table} ({','.join(eng_cols)}) VALUES ({placeholders})"

            inserted = 0
            for row in rows:
                if all(row[i] is None for i in col_indices if i is not None):
                    continue
                vals = []
                for i in col_indices:
                    vals.append(None if (i is None or i >= len(row)) else _val(row[i]))
                con.execute(sql_ins, vals)
                inserted += 1
            counts[table] = inserted
        wb.close()
        _create_indexes(con)
        con.commit()
        return {"ok": True, "counts": counts}
    except Exception as e:
        con.rollback()
        raise HTTPException(status_code=400, detail=f"엑셀 가져오기 오류: {e}")
    finally:
        con.close()


# ── POST /stddict/restore ────────────────────────────────────────────────────
# 브라우저가 vendor/std.sqlite(시드) bytes 를 전송 → 시스템 DB(aerm_storage)의
# 표준사전 테이블만 교체한다 (전체 파일 교체가 아니라 word/domain/term 만 →
# 시스템 DB의 다른 기능 테이블은 보존). 최초 초기화·시드 복원 양쪽에서 사용.
@router.post("/restore")
async def restore(file: UploadFile = File(...)):
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="빈 파일")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(suffix=".sqlite", dir=str(DATA_DIR))
    tmp = Path(tmp_path)
    con = None
    try:
        with open(fd, "wb") as f:
            f.write(data)
        con = _conn()
        _create_schema(con)
        con.execute("ATTACH DATABASE ? AS seed", [str(tmp)])
        try:
            for table, cols in STD_COLS.items():
                seed_has = con.execute(
                    "SELECT 1 FROM seed.sqlite_master WHERE type='table' AND name=?",
                    [table],
                ).fetchone()
                con.execute(f"DELETE FROM {table}")
                if seed_has:
                    collist = "id," + ",".join(cols)
                    con.execute(f"INSERT INTO {table} ({collist}) SELECT {collist} FROM seed.{table}")
            _create_indexes(con)
            con.commit()
        finally:
            con.execute("DETACH DATABASE seed")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"복원 실패: {e}")
    finally:
        if con is not None:
            con.close()
        if tmp.exists():
            tmp.unlink()

    return {"ok": True}


# ── GET /stddict/export ──────────────────────────────────────────────────────
@router.get("/export")
def export_db():
    _ensure_db()
    data = SYSTEM_DB_FILE.read_bytes()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/octet-stream",
        headers={"Content-Disposition": "attachment; filename=aerm_storage.db"},
    )
