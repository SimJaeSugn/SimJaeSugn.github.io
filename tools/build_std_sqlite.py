# tools/build_std_sqlite.py  (개발 전용 · 배포 제외)
# 실행: PYTHONIOENCODING=utf-8 python tools/build_std_sqlite.py
# 입력: docs/std-all-20260602-2151.xlsx
# 출력: vendor/std.sqlite
# 의존성: openpyxl (pip install openpyxl)

import sqlite3, datetime, sys, os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("오류: openpyxl 설치 필요 → pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── 경로 ──────────────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent.parent   # 프로젝트 루트
SRC  = BASE / "docs" / "std-all-20260602-2151.xlsx"
OUT  = BASE / "vendor" / "std.sqlite"

if not SRC.exists():
    print(f"오류: 입력 파일 없음 → {SRC}", file=sys.stderr)
    sys.exit(1)

# ── 시트 → 테이블 매핑 (헤더 한글 → 영문 컬럼명) ──────────────────
# 형식: { 시트명: (테이블명, [(한글헤더, 영문컬럼명), ...]) }
SHEETS = {
    "표준단어": ("word", [
        ("표준단어명",  "name"),
        ("영문약어",    "abbr"),
        ("영문전체명",  "full_name"),
        ("설명",        "descr"),
        ("도메인분류",  "domain_class"),
        ("형식단어",    "format_word"),
        ("이음동의어",  "synonym"),
        ("금칙어",      "forbidden"),
        ("개정구분",    "revision"),
        ("조직구분",    "org"),
        ("승인여부",    "approved"),
        ("등록자",      "reg_user"),
        ("등록일시",    "reg_at"),
        ("수정자",      "upd_user"),
        ("수정일시",    "upd_at"),
    ]),
    "표준도메인": ("domain", [
        ("도메인분류명", "class_name"),
        ("도메인명",     "name"),
        ("도메인그룹명", "group_name"),
        ("설명",         "descr"),
        ("데이터타입",   "data_type"),
        ("길이",         "len"),
        ("소수점",       "scale"),
        ("저장형식",     "store_fmt"),
        ("표현형식",     "disp_fmt"),
        ("단위",         "unit"),
        ("허용값",       "allowed"),
        ("개정구분",     "revision"),
        ("조직구분",     "org"),
        ("승인여부",     "approved"),
        ("등록자",       "reg_user"),
        ("등록일시",     "reg_at"),
    ]),
    "표준용어": ("term", [
        ("표준용어명",     "name"),
        ("영문약어",       "abbr"),
        ("설명",           "descr"),
        ("도메인명",       "domain_name"),
        ("허용값",         "allowed"),
        ("저장형식",       "store_fmt"),
        ("표현형식",       "disp_fmt"),
        ("행정표준코드명", "gov_code_name"),
        ("소관기관명",     "gov_org"),
        ("이음동의어",     "synonym"),
        ("개정구분",       "revision"),
        ("조직구분",       "org"),
        ("승인여부",       "approved"),
        ("등록자",         "reg_user"),
        ("등록일시",       "reg_at"),
    ]),
}

# ── 날짜 값 → 문자열 변환 (YYYY-MM-DD HH:MM:SS) ───────────────────
def _fmt_dt(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if v is None:
        return None
    return str(v)

# ── 셀 값 정규화 ──────────────────────────────────────────────────
def _val(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return _fmt_dt(v)
    if v is None:
        return None
    return str(v) if not isinstance(v, (int, float)) else v

# ── 메인 빌드 ─────────────────────────────────────────────────────
def build():
    print(f"입력: {SRC}")
    print(f"출력: {OUT}")

    # 기존 파일 삭제 후 재생성
    if OUT.exists():
        OUT.unlink()

    con = sqlite3.connect(str(OUT))
    cur = con.cursor()

    # WAL 모드 → 빌드 성능 향상
    cur.execute("PRAGMA journal_mode=WAL")

    wb = openpyxl.load_workbook(str(SRC), read_only=True, data_only=True)

    for sheet_name, (table, col_map) in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"경고: 시트 '{sheet_name}' 없음 — 건너뜀", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        eng_cols = [eng for _, eng in col_map]
        han_cols  = [han for han, _ in col_map]

        # ── CREATE TABLE ─────────────────────────────────────────
        col_defs = "id INTEGER PRIMARY KEY, " + ", ".join(f"{c} TEXT" for c in eng_cols)
        cur.execute(f"CREATE TABLE IF NOT EXISTS {table} ({col_defs})")

        # ── 헤더 행 인덱스 매핑 ──────────────────────────────────
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            print(f"경고: 시트 '{sheet_name}' 헤더 없음", file=sys.stderr)
            continue

        # 한글 헤더 → 컬럼 인덱스 맵
        header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        col_indices = []
        for han, eng in col_map:
            idx = header_idx.get(han)
            if idx is None:
                print(f"  경고: '{sheet_name}' 시트에 헤더 '{han}' 없음", file=sys.stderr)
                col_indices.append(None)
            else:
                col_indices.append(idx)

        # ── INSERT ───────────────────────────────────────────────
        placeholders = ", ".join("?" * len(eng_cols))
        sql_ins = f"INSERT INTO {table} ({', '.join(eng_cols)}) VALUES ({placeholders})"

        inserted = 0
        for row in rows:
            # 모든 값이 None인 빈 행 건너뜀
            if all(row[i] is None for i in col_indices if i is not None):
                continue
            vals = []
            for i in col_indices:
                if i is None or i >= len(row):
                    vals.append(None)
                else:
                    vals.append(_val(row[i]))
            cur.execute(sql_ins, vals)
            inserted += 1

        print(f"  {table}: {inserted}행 삽입")

    wb.close()

    # ── 인덱스 생성 ──────────────────────────────────────────────
    cur.execute("CREATE INDEX IF NOT EXISTS idx_word_name   ON word(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_word_abbr   ON word(abbr)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_domain_name ON domain(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_term_name   ON term(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_term_abbr   ON term(abbr)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_term_domain ON term(domain_name)")

    con.commit()

    # ── VACUUM ───────────────────────────────────────────────────
    cur.execute("PRAGMA journal_mode=DELETE")
    cur.close()
    con.commit()
    con.execute("VACUUM")
    con.commit()
    con.close()

    size_kb = OUT.stat().st_size // 1024
    print(f"\n완료: {OUT}  ({size_kb} KB)")

if __name__ == "__main__":
    build()
